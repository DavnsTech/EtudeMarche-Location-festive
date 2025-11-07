"""
Excel report generator for market study data
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os
from typing import Optional, Dict, Any

class MarketStudyExcelReport:
    def __init__(self, 
                 competitor_data_file: str = 'data/competitor_research.xlsx',
                 market_data_file: str = 'data/market_overview.xlsx'):
        """
        Initializes the Excel report generator.

        Args:
            competitor_data_file (str): Path to the competitor research Excel file.
            market_data_file (str): Path to the market overview Excel file.
        """
        self.wb = Workbook()
        self.competitor_data_file = competitor_data_file
        self.market_data_file = market_data_file
        self.reports_dir: str = 'reports'
        self.output_filename: str = os.path.join(self.reports_dir, 'Location_Festive_Niort_Market_Study_Report.xlsx')

        # Ensure reports directory exists
        if not os.path.exists(self.reports_dir):
            try:
                os.makedirs(self.reports_dir)
                print(f"Created directory: {self.reports_dir}")
            except OSError as e:
                print(f"Error creating directory {self.reports_dir}: {e}")

    def _style_header(self, ws, title: str, start_row: int = 1, start_col: int = 1) -> None:
        """Helper to style a title cell."""
        cell = ws.cell(row=start_row, column=start_col, value=title)
        cell.font = Font(size=18, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Adjust merge based on expected content width, assuming 5 columns for this title
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 4)

    def _style_section_title(self, ws, title: str, row: int, col: int = 1) -> int:
        """Helper to style section titles."""
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = Font(size=14, bold=True, color="000000")
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 4) # Assuming section titles span 5 columns
        return row + 2 # Return next row, with a blank row in between

    def _apply_bold_font(self, cell):
        """Applies bold font style to a cell."""
        cell.font = Font(bold=True)

    def _load_excel_data(self, filepath: str) -> Optional[pd.DataFrame]:
        """Loads data from an Excel file, handling potential errors."""
        if not os.path.exists(filepath):
            print(f"Error: Data file not found at {filepath}")
            return None
        try:
            df = pd.read_excel(filepath)
            return df
        except Exception as e:
            print(f"Error reading Excel file {filepath}: {e}")
            return None

    def create_executive_summary_sheet(self) -> None:
        """Creates the 'Résumé Exécutif' sheet."""
        ws = self.wb.active
        ws.title = "Résumé Exécutif"
        
        # Add main title
        self._style_header(ws, "ÉTUDE DE MARCHE - LOCATION FESTIVE NIORT")
        
        current_row = 4 # Start after title and some spacing

        # Market Overview Summary
        current_row = self._style_section_title(ws, "Aperçu du Marché", current_row)
        market_df = self._load_excel_data(self.market_data_file)
        if market_df is not None:
            market_df.columns = ['Category', 'Details'] # Ensure column names are consistent
            for r_idx, row in enumerate(dataframe_to_rows(market_df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)
                    if c_idx == 0: # Category column
                        self._apply_bold_font(cell)
                    cell.alignment = Alignment(wrapText=True, vertical="top")
            current_row += len(market_df) + 1 # Add space for the data and next section

        # Competitor Analysis Summary
        current_row = self._style_section_title(ws, "Analyse Concurrentielle", current_row)
        competitor_df = self._load_excel_data(self.competitor_data_file)
        if competitor_df is not None:
            # Basic stats from competitor analysis
            total_competitors = len(competitor_df)
            avg_strengths = competitor_df['Strengths'].dropna().apply(lambda x: len(str(x).split(',')) if pd.notnull(x) and str(x).strip() else 0).mean()
            avg_weaknesses = competitor_df['Weaknesses'].dropna().apply(lambda x: len(str(x).split(',')) if pd.notnull(x) and str(x).strip() else 0).mean()
            market_pos_counts = competitor_df['Market Position'].value_counts()

            summary_content = [
                ("Total Competitors", total_competitors),
                ("Average Strengths per Competitor", f"{avg_strengths:.2f}" if pd.notna(avg_strengths) else "N/A"),
                ("Average Weaknesses per Competitor", f"{avg_weaknesses:.2f}" if pd.notna(avg_weaknesses) else "N/A")
            ]
            
            for i, (metric, value) in enumerate(summary_content):
                ws.cell(row=current_row + i*2, column=1, value=metric).font = Font(bold=True)
                ws.cell(row=current_row + i*2, column=2, value=value)
            
            # Market Position Counts
            current_row += len(summary_content) * 2
            ws.cell(row=current_row, column=1, value="Market Position Distribution:").font = Font(bold=True)
            current_row += 1
            for position, count in market_pos_counts.items():
                ws.cell(row=current_row, column=1, value=f"- {position}")
                ws.cell(row=current_row, column=2, value=count)
                current_row += 1
            current_row += 1 # Space for next section

        # Add a placeholder for future insights
        current_row = self._style_section_title(ws, "Principales Observations et Recommandations", current_row)
        ws.cell(row=current_row, column=1, value="Placeholder for key insights and strategic recommendations.").alignment = Alignment(italic=True)
        current_row += 2

        # Adjust column widths for better readability
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50 # Cap width

    def create_competitor_analysis_sheet(self) -> None:
        """Creates the 'Analyse Concurrentielle' sheet."""
        ws = self.wb.create_sheet("Analyse Concurrentielle")
        
        self._style_header(ws, "ANALYSE DE LA CONCURRENCE", start_col=1)
        current_row = 4

        competitor_df = self._load_excel_data(self.competitor_data_file)
        if competitor_df is None:
            ws.cell(row=current_row, column=1, value="Données concurrentielles non disponibles.").font = Font(color="FF0000")
            return

        # Display raw competitor data
        current_row = self._style_section_title(ws, "Données Brutes des Concurrents", current_row)
        for r_idx, row in enumerate(dataframe_to_rows(competitor_df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)
                if r_idx == 0: # Header row
                    self._apply_bold_font(cell)
                cell.alignment = Alignment(wrapText=True, vertical="top")
        
        current_row += len(competitor_df) + 2 # Add space for data and next section

        # --- Analysis Section: Strengths vs Weaknesses ---
        current_row = self._style_section_title(ws, "Forces vs Faiblesses", current_row)
        
        # Calculate Strengths and Weaknesses counts
        competitor_df['Num_Strengths'] = competitor_df['Strengths'].dropna().apply(
            lambda x: len(str(x).split(',')) if pd.notnull(x) and str(x).strip() else 0
        )
        competitor_df['Num_Weaknesses'] = competitor_df['Weaknesses'].dropna().apply(
            lambda x: len(str(x).split(',')) if pd.notnull(x) and str(x).strip() else 0
        )

        # Create a DataFrame for this analysis section
        analysis_df = competitor_df[['Competitor', 'Num_Strengths', 'Num_Weaknesses']].copy()
        
        # Add analysis data to sheet
        for r_idx, row in enumerate(dataframe_to_rows(analysis_df, index=False, header=True)):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)
                if r_idx == 0: # Header row
                    self._apply_bold_font(cell)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        current_row += len(analysis_df) + 1

        # Add a bar chart for Strengths vs Weaknesses
        try:
            chart = BarChart()
            chart.title = "Nombre de Forces vs Faiblesses par Concurrent"
            chart.style = 10
            chart.y_axis.title = 'Nombre'
            chart.x_axis.title = 'Concurrent'
            chart.legend = None # No legend needed as Competitor is on X-axis

            data = Reference(ws, min_col=2, min_row=current_row - len(analysis_df), max_row=current_row - 1)
            cats = Reference(ws, min_col=1, min_row=current_row - len(analysis_df), max_row=current_row - 1)
            
            # Create separate references for strengths and weaknesses if needed for stacked/grouped
            # For simple comparison, plotting both as separate bars is good
            strengths_ref = Reference(ws, min_col=2, min_row=current_row - len(analysis_df) + 1, max_row=current_row - 1)
            weaknesses_ref = Reference(ws, min_col=3, min_row=current_row - len(analysis_df) + 1, max_row=current_row - 1)
            
            chart.add_data(strengths_ref, titles_from_data=True)
            chart.add_data(weaknesses_ref, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, "E7") # Position the chart
            current_row += 15 # Adjust row counter for chart space
        except Exception as e:
            print(f"Error creating bar chart for competitor analysis: {e}")

        # --- Analysis Section: Market Position ---
        current_row = self._style_section_title(ws, "Positionnement sur le Marché", current_row)
        market_pos_counts = competitor_df['Market Position'].value_counts()

        if not market_pos_counts.empty:
            pos_data = pd.DataFrame({'Market Position': market_pos_counts.index, 'Count': market_pos_counts.values})
            for r_idx, row in enumerate(dataframe_to_rows(pos_data, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)
                    if r_idx == 0: # Header row
                        self._apply_bold_font(cell)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += len(pos_data) + 1

            # Add a pie chart for Market Position distribution
            try:
                chart_pie = PieChart()
                chart_pie.title = "Répartition du Positionnement sur le Marché"
                
                labels = Reference(ws, min_col=1, min_row=current_row - len(pos_data), max_row=current_row - 1)
                data_pie = Reference(ws, min_col=2, min_row=current_row - len(pos_data), max_row=current_row - 1)
                
                chart_pie.add_data(data_pie, titles_from_data=True)
                chart_pie.set_categories(labels)
                chart_pie.dataLabels = True
                chart_pie.dataLabels.position = 'best'
                chart_pie.dataLabels.showPercent = True
                chart_pie.dataLabels.showVal = False
                chart_pie.legend.position = XL_LEGEND_POSITION.RIGHT
                
                ws.add_chart(chart_pie, "E15") # Position the chart
                current_row += 15 # Adjust row counter for chart space
            except Exception as e:
                print(f"Error creating pie chart for market position: {e}")
        else:
            ws.cell(row=current_row, column=1, value="Aucune donnée de positionnement sur le marché disponible.").font = Font(color="FF0000")
            current_row += 2

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50

    def create_market_trends_sheet(self) -> None:
        """Creates the 'Tendances du Marché' sheet."""
        ws = self.wb.create_sheet("Tendances du Marché")
        
        self._style_header(ws, "TENDANCES DU MARCHÉ", start_col=1)
        current_row = 4

        market_df = self._load_excel_data(self.market_data_file)
        if market_df is None:
            ws.cell(row=current_row, column=1, value="Données de marché non disponibles.").font = Font(color="FF0000")
            return

        # Filter for Trends and Seasonality
        trends_data = market_df[market_df['Category'].isin(['Key Trends', 'Seasonal Peaks'])]
        
        if not trends_data.empty:
            # Extracting individual items for better display
            trends_list = []
            for index, row in trends_data.iterrows():
                category = row['Category']
                details_str = row['Details']
                if pd.notna(details_str):
                    items = [item.strip() for item in details_str.split(',')]
                    for item in items:
                        trends_list.append({'Category': category, 'Detail': item})
            
            trends_analysis_df = pd.DataFrame(trends_list)

            for r_idx, row in enumerate(dataframe_to_rows(trends_analysis_df, index=False, header=True)):
                for c_idx, value in enumerate(row):
                    cell = ws.cell(row=current_row + r_idx, column=c_idx + 1, value=value)
                    if r_idx == 0: # Header row
                        self._apply_bold_font(cell)
                    cell.alignment = Alignment(wrapText=True, vertical="top")
            current_row += len(trends_analysis_df) + 2
        else:
            ws.cell(row=current_row, column=1, value="Aucune donnée sur les tendances du marché trouvée.").font = Font(color="FF0000")
            current_row += 2

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width if adjusted_width < 50 else 50

    def save_report(self) -> Optional[str]:
        """Saves the generated Excel workbook."""
        try:
            self.wb.save(self.output_filename)
            print(f"Excel report saved successfully to: {self.output_filename}")
            return self.output_filename
        except Exception as e:
            print(f"Error saving Excel report: {e}")
            return None

    def generate_report(self) -> Optional[str]:
        """Generates all sheets for the Excel report."""
        print("Generating Excel report...")
        self.create_executive_summary_sheet()
        self.create_competitor_analysis_sheet()
        self.create_market_trends_sheet()
        return self.save_report()

